# 04_inhalte_ändern.py
# In diesem Beispiel geht es um eine Produkttabelle in welcher sich ein Fehler eingeschlichen hat
# Jemand hat Bei Knoblauch, Zitronen und Staudensellerie versehentlich den falschen Preis angegeben
# Die Aufgabe des Scripts ist es nun nach den Einträgen für Knoblauch Zitronen und Staudensellerie 
# Ausschau zu halten und die entsprechenden Preise an der richtigen Position zu ändern.
# Spalte A muss zutreffen: garlic, celery, lemons Spalte B soll mit neuen Werten ersetzt werden (cost per pound).

import openpyxl, os
os.chdir(os.path.dirname(__file__))

new_prices={'garlic' : 3.07, 'celery' : 1.19, 'lemon' : 1.27}

excel_file='.\\produktliste.xlsx'
new_excel_file='.\\korrigierte_produktliste.xlsx'
if os.path.exists(new_excel_file):
    os.remove(new_excel_file)

# Lade Exceltabelle und fokussiere auf erste Arbeitsmappe
excel_sheet=openpyxl.load_workbook(excel_file)
sheet_names=excel_sheet.sheetnames
active_sheet=excel_sheet[sheet_names[0]]
max_row=active_sheet.max_row

#Starte von Position A2:B2 bis Ende der Tabelle
for i in range(2,max_row):
    if str(active_sheet.cell(row=i, column=1).value.lower()) in new_prices.keys():
        active_sheet.cell(row=i, column=2, value=new_prices[active_sheet.cell(row=i, column=1).value.lower()])

excel_sheet.save(new_excel_file)