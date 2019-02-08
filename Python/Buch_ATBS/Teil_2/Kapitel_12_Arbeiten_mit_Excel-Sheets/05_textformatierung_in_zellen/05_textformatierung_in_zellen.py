# 05_textformatierung_in_zellen.py
# In diesem Beispielscript geht es darum die Formatierung der Zellen zu beeinflussen
# Aus textfile.txt wird gezählt welche wörter wie oft vorkommen um diese dann entsprechend
# ihrer Anzahl in anderer Schriftart/-grösse darzustellen

import openpyxl, os
from openpyxl.styles import Font, Alignment
os.chdir(os.path.dirname(__file__))

source_file='.\\textfile.txt'
target_file='.\\word_cloud.xlsx'
if os.path.exists(target_file):
    os.remove(target_file)

cleaned_cloud, word_cloud={}, {}

# Zähle vorkommen der einzelnen Wörter innerhalb des angegebenen Files
read_file=open(source_file, 'r', encoding='UTF-8')
source_file_content=read_file.read()
read_file.close()
for possible_word in source_file_content.split(' '):
    if possible_word.strip('(),.\n\t\'"').isalpha() and len(possible_word.strip('(),.\n\t\'"')) > 3:
        word_cloud.setdefault(possible_word.strip('(),.\n\t\'"'), 0)
        word_cloud[possible_word.strip('(),.\n\t\'"')]+=1

# Entferne alle Wörter die weniger als min_anzahl vorkommen
min_anzahl=7
for eintrag in word_cloud:
    if word_cloud[eintrag] > 12:
        cleaned_cloud.setdefault(eintrag, word_cloud[eintrag])

# Bereite Dokument zum schreiben vor
excel_sheet=openpyxl.Workbook()
active_sheet=excel_sheet.active
active_sheet.freeze_panes='A2'

# Diverse Möglichkeiten zur anpassung der Schrift sowie Ausrichtung und Aussehen des Textes mit dem Modul Font
active_sheet.title = 'Word-Cloud'
sheet_title='Wordcloud des Dokuments '+source_file
active_sheet.cell(row=1, column=1, value=sheet_title).font=Font(size=20, bold=True, underline='single', name='Times New Roman')
active_sheet.row_dimensions[1].height=22

countinue, row, colums=True, 2, 0
max_colums=8

# Setze Titel im und vom Abeitsblatt
merge_range='A1:'+str(openpyxl.utils.get_column_letter(max_colums)+'1')
active_sheet.merge_cells(merge_range)
max_column_with={}
while countinue:
    row+=1
    max_row_size=0
    for i in range(1, max_colums+1):
        if i+colums+1 > len(list(cleaned_cloud.keys())):
            countinue=False
            break
        content=list(cleaned_cloud.keys())[i+colums]
        max_column_with.setdefault(openpyxl.utils.get_column_letter(i), 0)
        # Merke Zahl für Spaltenbreite und höhe der Reihe
        if max_row_size < cleaned_cloud[content]:
            max_row_size=cleaned_cloud[content]
        if max_column_with[openpyxl.utils.get_column_letter(i)] < len(content):
            max_column_with[openpyxl.utils.get_column_letter(i)] = len(content)
        # Zelleneintrag mit der Schriftgrösse = Anzahl des Wortes
        active_sheet.cell(row=row, column=i, value=content).font=Font(size=cleaned_cloud[content])
        active_sheet.cell(row=row, column=i).alignment=Alignment(horizontal='center')
    # Seitze höhe der Reihe auf die Anzahl des meistvorkommenden Wortes (= Schriftgrösse)
    active_sheet.row_dimensions[row].height=max_row_size
    colums+=10

# Passe Spaltenbreite dem längsten Wort + Anzahl des vorkommenden Wortes // 2.2 an
position=1
for colum in max_column_with:
    active_sheet.column_dimensions[colum].width=max_column_with[colum]+cleaned_cloud[list(cleaned_cloud.keys())[position]]//2.2
    position+=1

excel_sheet.save(target_file)
