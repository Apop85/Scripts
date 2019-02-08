# 06_erstellen_von_diagrammen.py
# In diesem Beispiel geht es darum automatisch Diagramme in ein Excelsheet einzufügen.

import os, openpyxl
from openpyxl.styles import Font, Alignment
from math import pi, sin

os.chdir(os.path.dirname(__file__))

target_file='.\\excel_with_diagrams.xlsx'

excel_file=openpyxl.Workbook()
active_sheet=excel_file.active
active_sheet.title='Diagramm'
active_sheet.cell(row=1, column=1, value='Erstellen von Diagrammen mittels Python').font=Font(size=22, name='Times New Roman', underline='single', bold=True)
max_colums=10
active_sheet.merge_cells='A1'+':'+'A'+str(max_colums)

# Nehme Sinuskurvenwerte als Datengrundlage
dataset={}
for i in range(1,101):
    active_sheet.cell(row=i+2, column=1, value=sin(2*pi*500*i/10000))

# Definieren des Diagramms
max_rows=active_sheet.max_row
diagram_coordinates=openpyxl.chart.Reference(active_sheet, min_col=1, min_row=3, max_col=1, max_row=max_rows)
diagram_series=openpyxl.chart.Series(diagram_coordinates, title='Sinuskurve')
diagram_object=openpyxl.chart.LineChart()

# Füge Diagramm mit vorgegebenen Dimensionen ein.
diagram_object.append(diagram_series)
diagram_object.title='Sinuskurvendiagramm'
diagram_object.smooth
active_sheet.add_chart(diagram_object, 'C3')

excel_file.save(target_file)
