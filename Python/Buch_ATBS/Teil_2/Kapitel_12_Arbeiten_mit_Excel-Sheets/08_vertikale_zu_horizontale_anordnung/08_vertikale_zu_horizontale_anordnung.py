# 08_uebung_einträge_vertikale_zu_hotizontale_anordnung.py
# In dieser Übung geht es darum Einträge einer Tabelle die Vertikal geordnet 
# sind horizontal auszurichten

import openpyxl, os
os.chdir(os.path.dirname(__file__))

example_file='.\\example.xlsx'
target_file='.\\output.xlsx'

if os.path.exists(target_file):
    os.remove(target_file)

work_sheet=openpyxl.load_workbook(example_file)
active_sheet=work_sheet.active

# Ermittle maximale Anzahl Reihen und Spalten
max_rows=active_sheet.max_row
max_colums=active_sheet.max_column
max_colum_name=openpyxl.utils.get_column_letter(max_colums)

# Bereite zur Datenveratbeitung vor.
sheet_content={}
new_sheet=openpyxl.Workbook()
new_active_sheet=new_sheet.active
new_active_sheet.title='Sorting around'
# Daten auslesen und umsortieren
for i in range(1,max_rows+1):
    # Lese Daten aus Originaldatei aus.
    column_title=active_sheet.cell(row=i, column=1).value
    row_content=active_sheet['B'+str(i)+':'+max_colum_name+str(i)]
    new_active_sheet.cell(row=1, column=i, value=column_title)
    # Schreibe in neue Datei
    for content in row_content:
        counter=2
        for tuple_content in content:
            new_active_sheet.cell(row=counter, column=i, value=str(tuple_content.value))
            counter+=1

new_sheet.save(target_file)

